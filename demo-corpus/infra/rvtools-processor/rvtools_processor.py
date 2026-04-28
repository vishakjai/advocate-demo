import openpyxl
import pandas as pd
import sys
import os
import argparse
from pathlib import Path
import json
from datetime import datetime
import warnings
import ipaddress
import hashlib
import re

# Suppress pandas warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

def find_rvtools_files(directory="."):
    """Find RVTools files in directory"""
    files = []
    xlsx_files = list(Path(directory).glob("*.xlsx"))
    
    for file in xlsx_files:
        if file.name.startswith("~"):
            continue
        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            rvtools_sheets = {'vInfo', 'vHost', 'vCluster'}
            if any(sheet in wb.sheetnames for sheet in rvtools_sheets):
                files.append(file)
        except Exception:
            continue
    
    return files

def generate_output_filename(mode, input_files=None):
    """Generate output filename with timestamp"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    
    if mode == "both":
        return f"RVTools_Consolidated_Anonymized_{timestamp}.xlsx"
    elif mode == "consolidate":
        return f"RVTools_Combined_{timestamp}.xlsx"
    elif mode == "anonymize":
        return f"RVTools_Anonymized_{timestamp}.xlsx"
    elif mode == "deanonymize":
        return f"RVTools_Deanonymized_{timestamp}.xlsx"

def consolidate_rvtools(input_files, output_file):
    """Consolidate multiple RVTools files into one"""
    print(f"\nStarting consolidation of {len(input_files)} files...")
    consolidated_sheets = {}
    
    for file_path in input_files:
        print(f"Processing: {file_path.name}")
        try:
            wb = openpyxl.load_workbook(filename=file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                if sheet_name not in consolidated_sheets:
                    consolidated_sheets[sheet_name] = df
                else:
                    consolidated_sheets[sheet_name] = pd.concat([consolidated_sheets[sheet_name], df], 
                                                                ignore_index=True)
            wb.close()
        except Exception as e:
            print(f"Error processing {file_path.name}: {str(e)}")
            continue

    print(f"Creating consolidated file: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in consolidated_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("\nConsolidation Summary:")
    print(f"Input files processed: {len(input_files)}")
    print("Sheets consolidated:")
    for sheet_name, df in consolidated_sheets.items():
        print(f" - {sheet_name}: {len(df)} total rows")
    
    return output_file

class AnonymizationManager:
    """Manages anonymization mappings and transformations"""
    
    def __init__(self, filename_suffix=""):
        self.vm_counter = 1
        self.host_counter = 1
        self.cluster_counter = 1
        self.datacenter_counter = 1
        self.ip_mappings = {}
        self.name_mappings = {}
        self.reverse_mappings = {}
        self.vm_id_mappings = {}  # Map VM names to their VM IDs
        self.filename_suffix = filename_suffix
        # Track unique entities to avoid double counting
        self.unique_vms = set()
        self.unique_hosts = set()
        self.unique_clusters = set()
        self.unique_datacenters = set()
        
    def anonymize_vm_name_with_id(self, vm_name, vm_id=None, count_vm=True):
        """Anonymize VM name using VM ID when available, with filename suffix for duplicates"""
        if not vm_name or pd.isna(vm_name):
            return vm_name
            
        vm_name_str = str(vm_name)
        
        # Check if we already processed this VM name
        if vm_name_str in self.name_mappings:
            return self.name_mappings[vm_name_str]
        
        # Track unique VM for counting only if requested
        if count_vm:
            self.unique_vms.add(vm_name_str)
        
        # If we have a VM ID, use it as the anonymized name
        if vm_id and not pd.isna(vm_id):
            vm_id_str = str(vm_id)
            # Add filename suffix if this is from a consolidated file and there might be duplicates
            if self.filename_suffix and vm_id_str in self.vm_id_mappings:
                anon_name = f"{vm_id_str}_{self.filename_suffix}"
            else:
                anon_name = vm_id_str
            
            self.name_mappings[vm_name_str] = anon_name
            self.reverse_mappings[anon_name] = vm_name_str
            self.vm_id_mappings[vm_id_str] = anon_name
            return anon_name
        
        # Fallback to sequential naming if no VM ID
        anon_name = f"VM-{self.vm_counter:04d}"
        if self.filename_suffix:
            anon_name += f"_{self.filename_suffix}"
        self.name_mappings[vm_name_str] = anon_name
        self.reverse_mappings[anon_name] = vm_name_str
        self.vm_counter += 1
        return anon_name
    
    def anonymize_host_name(self, host_name, count_host=True):
        """Anonymize host name while maintaining consistency"""
        if not host_name or pd.isna(host_name):
            return host_name
            
        host_name_str = str(host_name)
        if host_name_str not in self.name_mappings:
            if count_host:
                self.unique_hosts.add(host_name_str)
            anon_name = f"HOST-{self.host_counter:04d}"
            self.name_mappings[host_name_str] = anon_name
            self.reverse_mappings[anon_name] = host_name_str
            self.host_counter += 1
        return self.name_mappings[host_name_str]
    
    def anonymize_cluster_name(self, cluster_name):
        """Anonymize cluster name while maintaining consistency"""
        if not cluster_name or pd.isna(cluster_name):
            return cluster_name
            
        cluster_name_str = str(cluster_name)
        if cluster_name_str not in self.name_mappings:
            self.unique_clusters.add(cluster_name_str)
            anon_name = f"CLUSTER-{self.cluster_counter:04d}"
            self.name_mappings[cluster_name_str] = anon_name
            self.reverse_mappings[anon_name] = cluster_name_str
            self.cluster_counter += 1
        return self.name_mappings[cluster_name_str]
    
    def anonymize_datacenter_name(self, dc_name):
        """Anonymize datacenter name while maintaining consistency"""
        if not dc_name or pd.isna(dc_name):
            return dc_name
            
        dc_name_str = str(dc_name)
        if dc_name_str not in self.name_mappings:
            self.unique_datacenters.add(dc_name_str)
            anon_name = f"DC-{self.datacenter_counter:04d}"
            self.name_mappings[dc_name_str] = anon_name
            self.reverse_mappings[anon_name] = dc_name_str
            self.datacenter_counter += 1
        return self.name_mappings[dc_name_str]
    
    def anonymize_ip_address(self, ip_str):
        """Anonymize IP address while preserving network structure"""
        if not ip_str or pd.isna(ip_str) or str(ip_str).strip() == '':
            return ip_str
            
        ip_str = str(ip_str).strip()
        
        # Handle multiple IPs separated by comma or semicolon
        if ',' in ip_str or ';' in ip_str:
            separator = ',' if ',' in ip_str else ';'
            ips = [ip.strip() for ip in ip_str.split(separator)]
            anonymized_ips = []
            for ip in ips:
                anon_ip = self.anonymize_single_ip(ip)
                if anon_ip:
                    anonymized_ips.append(anon_ip)
            return separator.join(anonymized_ips) if anonymized_ips else ip_str
        else:
            return self.anonymize_single_ip(ip_str) or ip_str
    
    def anonymize_single_ip(self, ip_str):
        """Anonymize a single IP address (IPv4 and IPv6)"""
        if not ip_str or ip_str in self.ip_mappings:
            return self.ip_mappings.get(ip_str, ip_str)
        
        try:
            # Try to parse as IPv4 first
            ip_obj = ipaddress.IPv4Address(ip_str)
            
            # Create a hash-based mapping that preserves network structure
            ip_int = int(ip_obj)
            
            # Use first 3 octets to determine network, anonymize last octet
            network_part = ip_int & 0xFFFFFF00  # Keep first 3 octets
            host_part = ip_int & 0xFF  # Last octet
            
            # Create consistent anonymization based on original network
            network_hash = hashlib.md5(str(network_part).encode()).hexdigest()[:4]
            network_id = int(network_hash, 16) % 254 + 1  # 1-254 range
            
            # Map to 10.x.x.x private range to avoid conflicts
            anon_network = 0x0A000000 | (network_id << 16) | ((network_part >> 8) & 0xFF00)
            anon_ip = anon_network | host_part
            
            anon_ip_str = str(ipaddress.IPv4Address(anon_ip))
            self.ip_mappings[ip_str] = anon_ip_str
            self.reverse_mappings[anon_ip_str] = ip_str
            
            return anon_ip_str
            
        except (ipaddress.AddressValueError, ValueError):
            try:
                # Try to parse as IPv6
                ip_obj = ipaddress.IPv6Address(ip_str)
                
                # For IPv6, create a hash-based anonymization
                ip_hash = hashlib.md5(ip_str.encode()).hexdigest()
                # Create a valid IPv6 address in the documentation range (2001:db8::/32)
                anon_ipv6 = f"2001:db8:{ip_hash[:4]}:{ip_hash[4:8]}:{ip_hash[8:12]}:{ip_hash[12:16]}:{ip_hash[16:20]}:{ip_hash[20:24]}"
                
                self.ip_mappings[ip_str] = anon_ipv6
                self.reverse_mappings[anon_ipv6] = ip_str
                
                return anon_ipv6
                
            except (ipaddress.AddressValueError, ValueError):
                # If not a valid IP, just return as-is or create generic mapping
                if re.match(r'^\d+\.\d+\.\d+\.\d+$', ip_str):
                    # Looks like IPv4 but invalid, anonymize anyway
                    anon_ip = f"10.0.0.{len(self.ip_mappings) % 254 + 1}"
                    self.ip_mappings[ip_str] = anon_ip
                    self.reverse_mappings[anon_ip] = ip_str
                    return anon_ip
                return ip_str
    
    def anonymize_generic_field(self, value, field_type="ANON"):
        """Anonymize generic fields"""
        if not value or pd.isna(value):
            return value
            
        value_str = str(value)
        if value_str not in self.name_mappings:
            anon_value = f"{field_type}-{len(self.name_mappings):04d}"
            self.name_mappings[value_str] = anon_value
            self.reverse_mappings[anon_value] = value_str
        return self.name_mappings[value_str]
    
    def get_mappings(self):
        """Return all mappings for saving"""
        return {
            'reverse_mappings': self.reverse_mappings,
            'ip_mappings': self.ip_mappings,
            'name_mappings': self.name_mappings
        }

def anonymize_rvtools(input_file, output_file):
    """Anonymize RVTools file preserving network structure and relationships"""
    print(f"\nStarting anonymization of {input_file}...")
    
    # Extract filename for suffix in case of consolidation
    filename_base = Path(input_file).stem.replace('RVTools_export_all_', '').replace('RVTools_Combined_', '')
    anon_manager = AnonymizationManager(filename_base)
    
    try:
        # Validate input file exists
        if not Path(input_file).exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")
        
        # Read the Excel file
        wb = openpyxl.load_workbook(filename=input_file)
        
        # Validate it's an RVTools file
        rvtools_sheets = {'vInfo', 'vHost', 'vCluster'}
        if not any(sheet in wb.sheetnames for sheet in rvtools_sheets):
            print("Warning: This doesn't appear to be an RVTools file (missing expected sheets)")
        
        # Process sheets in specific order to ensure proper counting
        priority_sheets = ['vHost', 'vCluster', 'vInfo']  # Process these first for accurate counting
        remaining_sheets = [s for s in wb.sheetnames if s not in priority_sheets]
        sheet_order = priority_sheets + remaining_sheets
        
        for sheet_name in sheet_order:
            if sheet_name not in wb.sheetnames:
                continue
            print(f"Processing sheet: {sheet_name}")
            sheet = wb[sheet_name]
            
            if sheet.max_row <= 1:  # Skip empty sheets
                continue
                
            # Get headers
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # Find VM ID column if it exists
            vm_id_col_idx = None
            if 'VM ID' in headers:
                vm_id_col_idx = headers.index('VM ID')
            
            # Process each row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                # Get VM ID for this row if available
                vm_id = None
                if vm_id_col_idx is not None and vm_id_col_idx < len(row):
                    vm_id = row[vm_id_col_idx].value
                
                for col_idx, cell in enumerate(row):
                    if col_idx >= len(headers):
                        continue
                        
                    header = headers[col_idx]
                    original_value = cell.value
                    
                    if not original_value or pd.isna(original_value) or str(original_value).strip() == '':
                        continue
                    
                    # Skip if it's just a number (like port numbers, IDs that shouldn't be anonymized)
                    if header in ['Port', 'VLAN', 'VLAN ID', 'CPU', 'Memory', 'Disk Size', 'Used Space', 'Free Space'] and str(original_value).isdigit():
                        continue
                    
                    # Anonymize based on field type and sheet
                    if header in ['VM', 'VM Name', 'Name'] and sheet_name in ['vInfo', 'vSnapshot', 'vCPU', 'vMemory', 'vDisk', 'vPartition', 'vNetwork', 'vFloppy', 'vCD']:
                        # Only count VMs from vInfo sheet to avoid duplicates
                        count_vm = (sheet_name == 'vInfo')
                        cell.value = anon_manager.anonymize_vm_name_with_id(original_value, vm_id, count_vm)
                    
                    elif header in ['Host', 'ESX Host', 'Host Name', 'Hosts']:
                        # Only count hosts from vHost sheet to avoid duplicates
                        count_host = (sheet_name == 'vHost')
                        cell.value = anon_manager.anonymize_host_name(original_value, count_host)
                    
                    elif header in ['Cluster', 'Cluster Name']:
                        cell.value = anon_manager.anonymize_cluster_name(original_value)
                    
                    elif header in ['Datacenter', 'Datacenter Name']:
                        cell.value = anon_manager.anonymize_datacenter_name(original_value)
                    
                    # IP Address fields (IPv4 and IPv6)
                    elif header in ['IP Address', 'Primary IP Address', 'IPv4 Address', 'IPv6 Address', 'IP Addresses', 'Address']:
                        cell.value = anon_manager.anonymize_ip_address(original_value)
                    
                    # DNS related fields
                    elif header in ['DNS Name', 'FQDN', 'DNS Server', 'DNS Servers']:
                        if '.' in str(original_value):
                            parts = str(original_value).split('.')
                            anon_hostname = anon_manager.anonymize_generic_field(parts[0], "HOST")
                            cell.value = f"{anon_hostname}.anon.local"
                        else:
                            cell.value = anon_manager.anonymize_generic_field(original_value, "DNS")
                    
                    # MAC Address
                    elif header in ['Mac Address', 'MAC Address']:
                        mac_hash = hashlib.md5(str(original_value).encode()).hexdigest()[:12]
                        formatted_mac = ':'.join([mac_hash[i:i+2] for i in range(0, 12, 2)])
                        cell.value = formatted_mac
                        anon_manager.reverse_mappings[formatted_mac] = str(original_value)
                    
                    # Path and directory fields
                    elif header in ['Folder', 'Path', 'Log directory', 'Snapshot directory', 'Suspend directory', 'Resource Pool path']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "PATH")
                    
                    # Disk related fields
                    elif header in ['Disk', 'Disk Path', 'Filename', 'File', 'Display Name'] and sheet_name in ['vDisk', 'vPartition', 'vMultiPath']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "DISK")
                    
                    # Internal sort columns (often contain server names)
                    elif header in ['Internal Sort Column']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "SORT")
                    
                    # Annotation and description fields
                    elif header in ['Annotation', 'Notes', 'Description', 'Message']:
                        cell.value = "ANONYMIZED_TEXT"
                    
                    # vCenter and SDK servers
                    elif header in ['VI SDK Server', 'vCenter']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "VCENTER")
                    
                    # Resource pools
                    elif header in ['Resource Pool', 'Resource pool']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "POOL")
                    
                    # Datastores
                    elif header in ['Datastore', 'Datastore Name']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "DS")
                    
                    # Network and switch related
                    elif header in ['Network', 'Network Name', 'Switch', 'vSwitch']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "NET")
                    
                    # URLs
                    elif header in ['URL']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "URL")
                    
                    # vHealth specific - Name field often contains server names
                    elif header in ['Name'] and sheet_name == 'vHealth':
                        cell.value = anon_manager.anonymize_generic_field(original_value, "HEALTH")
                    
                    # Additional fields that might contain sensitive information
                    elif header in ['Guest OS', 'Guest OS Full Name'] and 'hostname' in str(original_value).lower():
                        # Only anonymize if it looks like it contains hostname info
                        cell.value = anon_manager.anonymize_generic_field(original_value, "OS")
                    
                    elif header in ['Template', 'Template Name']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "TEMPLATE")
                    
                    elif header in ['VM UUID', 'Instance UUID'] and len(str(original_value)) > 10:
                        # Anonymize UUIDs but keep format
                        uuid_hash = hashlib.md5(str(original_value).encode()).hexdigest()
                        if '-' in str(original_value):
                            # Standard UUID format
                            cell.value = f"{uuid_hash[:8]}-{uuid_hash[8:12]}-{uuid_hash[12:16]}-{uuid_hash[16:20]}-{uuid_hash[20:32]}"
                        else:
                            cell.value = uuid_hash
                        anon_manager.reverse_mappings[cell.value] = str(original_value)
                    
                    elif header in ['Config File', 'VMX File']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "CONFIG")
                    
                    elif header in ['Snapshot Name', 'Snapshot Description']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "SNAPSHOT")
                    
                    elif header in ['License Key', 'Serial Number']:
                        cell.value = "ANONYMIZED_LICENSE"
                    
                    elif header in ['User', 'Owner', 'Created By']:
                        cell.value = anon_manager.anonymize_generic_field(original_value, "USER")
        
        # Save anonymized file
        wb.save(output_file)
        
        # Save mapping file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        mapping_file = f"mapping_{timestamp}.json"
        
        with open(mapping_file, 'w') as f:
            json.dump(anon_manager.get_mappings(), f, indent=2)
        
        print(f"✓ Created anonymized file: {output_file}")
        print(f"✓ Created mapping file: {mapping_file}")
        
        # Print detailed summary
        print("\n Anonymization Summary:")
        print(f"  • VMs: {len(anon_manager.unique_vms)}")
        print(f"  • Hosts: {len(anon_manager.unique_hosts)}")
        print(f"  • Clusters: {len(anon_manager.unique_clusters)}")
        print(f"  • Datacenters: {len(anon_manager.unique_datacenters)}")
        print(f"  • IP Addresses: {len(anon_manager.ip_mappings)}")
        print(f"  • Total mappings: {len(anon_manager.name_mappings)}")
        
        # Show IPv6 count if any
        ipv6_count = sum(1 for ip in anon_manager.ip_mappings.keys() if ':' in ip)
        if ipv6_count > 0:
            print(f"  • IPv6 Addresses: {ipv6_count}")
            print(f"  • IPv4 Addresses: {len(anon_manager.ip_mappings) - ipv6_count}")
        
        return output_file, mapping_file
        
    except Exception as e:
        print(f"Error during anonymization: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None

def deanonymize_rvtools(input_file, mapping_file, output_file):
    """Restore original names using mapping file"""
    print(f"\nStarting deanonymization...")
    
    try:
        # Load mapping
        with open(mapping_file, 'r') as f:
            mapping_data = json.load(f)
        
        # Extract reverse mappings
        reverse_mappings = mapping_data.get('reverse_mappings', {})
        
        wb = openpyxl.load_workbook(filename=input_file)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            if sheet.max_row <= 1:  # Skip empty sheets
                continue
            
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.value and str(cell.value) in reverse_mappings:
                        cell.value = reverse_mappings[str(cell.value)]
        
        wb.save(output_file)
        print(f"✓ Created deanonymized file: {output_file}")
        return output_file
        
    except Exception as e:
        print(f"Error during deanonymization: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def process_rvtools(args):
    """Main processing function supporting all modes"""
    input_files = []
    
    if not hasattr(args, 'input_files') or not args.input_files:
        input_files = find_rvtools_files()
        if not input_files:
            raise ValueError("No RVTools files found in current directory")
        print(f"Found {len(input_files)} RVTools files: {[f.name for f in input_files]}")
    else:
        input_files = [Path(f) for f in args.input_files]

    if not hasattr(args, 'output') or not args.output:
        args.output = generate_output_filename(args.mode, input_files)

    if args.mode == "consolidate":
        return consolidate_rvtools(input_files, args.output)
    elif args.mode == "anonymize":
        return anonymize_rvtools(input_files[0], args.output)
    elif args.mode == "deanonymize":
        if not args.mapping:
            raise ValueError("Mapping file required for deanonymization")
        return deanonymize_rvtools(input_files[0], args.mapping, args.output)
    elif args.mode == "both":
        temp_file = f"temp_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        consolidated = consolidate_rvtools(input_files, temp_file)
        result, mapping = anonymize_rvtools(consolidated, args.output)
        os.remove(temp_file)
        return result

def main():
    parser = argparse.ArgumentParser(description="RVTools Processing Tool")
    subparsers = parser.add_subparsers(dest='mode')
    
    # Consolidate parser
    parser_consolidate = subparsers.add_parser('consolidate',
                                             help='Consolidate multiple RVTools files')
    parser_consolidate.add_argument('input_files', nargs='*',
                                  help='Input files (optional, will find all RVTools files if not specified)')
    parser_consolidate.add_argument('-o', '--output',
                                  help='Output filename (optional)')

    # Anonymize parser
    parser_anonymize = subparsers.add_parser('anonymize',
                                           help='Anonymize RVTools file')
    parser_anonymize.add_argument('input_files', nargs='*',
                                help='Input file (optional, will find RVTools files if not specified)')
    parser_anonymize.add_argument('-o', '--output',
                                help='Output filename (optional)')
    parser_anonymize.add_argument('--dry-run', action='store_true',
                                help='Preview what would be anonymized without creating files')

    # Deanonymize parser
    parser_deanonymize = subparsers.add_parser('deanonymize',
                                             help='Deanonymize RVTools file')
    parser_deanonymize.add_argument('input_files', nargs='*',
                                  help='Input anonymized file')
    parser_deanonymize.add_argument('-m', '--mapping', required=True,
                                  help='Mapping file from anonymization')
    parser_deanonymize.add_argument('-o', '--output',
                                  help='Output filename (optional)')

    # Both consolidate and anonymize
    parser_both = subparsers.add_parser('both',
                                       help='Consolidate and anonymize RVTools files')
    parser_both.add_argument('input_files', nargs='*',
                           help='Input files (optional, will find all RVTools files if not specified)')
    parser_both.add_argument('-o', '--output',
                           help='Output filename (optional)')

    args = parser.parse_args()
    
    if not args.mode:
        parser.print_help()
        sys.exit(1)
    
    try:
        process_rvtools(args)
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
 
